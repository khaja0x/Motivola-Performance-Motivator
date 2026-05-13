from typing import Any, List
import uuid
from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.core.database import get_db
from app.api.deps import get_current_user, get_current_admin, get_current_supervisor, get_tenant_db
from app.models.models import Staff, Store, User, CommissionReport, Company
from app.schemas.staff import StaffCreate, StaffUpdate, StaffOut
from datetime import datetime
import shutil
from pathlib import Path

from fastapi.responses import StreamingResponse
import io
import pandas as pd
from sqlalchemy import select
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
import re

router = APIRouter()

@router.get("/", response_model=List[StaffOut])
async def list_staff(
    month: int | None = None,
    year: int | None = None,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_supervisor)
) -> Any:
    # 1. Fetch all staff for the company
    result = await db.execute(
        select(Staff).where(Staff.company_id == current_user.company_id)
    )
    staff_members = result.scalars().all()
    
    # 2. Extract targets for month/year (needed if we still rely heavily on CommissionReport, but instead we will aggregate from SalesHeader)
    now = datetime.utcnow()
    target_month = month if month else now.month
    target_year = year if year else now.year
    
    # We aggregate directly from Sales tables to ensure real-time accuracy for dashboards
    from app.models.models import SalesHeader, SalesLine
    from sqlalchemy import func
    
    sales_query = select(
        SalesHeader.staff_id,
        func.sum(SalesLine.total_amount).label("total_amount"),
        func.sum(SalesLine.quantity).label("total_quantity")
    ).join(
        SalesLine, SalesLine.sales_header_id == SalesHeader.id
    ).where(
        (SalesHeader.company_id == current_user.company_id) &
        (SalesHeader.status != 'Cancelled') &
        (func.extract('month', SalesHeader.order_date) == target_month) &
        (func.extract('year', SalesHeader.order_date) == target_year)
    ).group_by(SalesHeader.staff_id)
    
    sales_result = await db.execute(sales_query)
    actuals = {row.staff_id: row for row in sales_result.all()}

    # Also grab latest commission reports just for preserving the "commission" field
    report_result = await db.execute(
        select(CommissionReport).where(
            (CommissionReport.company_id == current_user.company_id) &
            (CommissionReport.month == target_month) &
            (CommissionReport.year == target_year)
        )
    )
    reports = {r.staff_id: r for r in report_result.scalars().all()}
    
    # 3. Calculate ranks based on total sales
    sorted_actuals = sorted(actuals.values(), key=lambda x: x.total_amount or 0, reverse=True)
    ranks = {row.staff_id: i + 1 for i, row in enumerate(sorted_actuals)}

    # 4. Enforce performance data into staff objects for response
    for s in staff_members:
        report = reports.get(s.staff_id)
        stat = actuals.get(s.staff_id)
        
        s.total_sales = float(stat.total_amount) if stat and stat.total_amount else 0.0
        s.total_sales_quantity = float(stat.total_quantity) if stat and stat.total_quantity else 0.0
        s.commission = report.calculated_commission if report else 0.0
        s.rank = ranks.get(s.staff_id)

    return staff_members

@router.post("/", response_model=StaffOut)
async def create_staff(
    staff_in: StaffCreate,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_admin)
) -> Any:
    from app.core.security import get_password_hash
    
    # 1. Ensure email exists and is unique within the company
    if not staff_in.email:
        raise HTTPException(status_code=400, detail="Email is required to create a staff login account")
        
    # Check if email already exists in users table (globally unique)
    user_check = await db.execute(select(User).where(User.email == staff_in.email))
    if user_check.scalars().first():
        raise HTTPException(status_code=400, detail="Email already registered in users")
        
    # Check if email already exists in staff table FOR THIS COMPANY ONLY
    staff_check = await db.execute(
        select(Staff).where(
            (Staff.email == staff_in.email) & 
            (Staff.company_id == current_user.company_id)
        )
    )
    if staff_check.scalars().first():
        raise HTTPException(status_code=400, detail="Email already registered in staff for your company")

    # Verify store belongs to company
    if staff_in.store_id is not None:
        store_result = await db.execute(
            select(Store).where(
                (Store.store_id == staff_in.store_id) & (Store.company_id == current_user.company_id)
            )
        )
        if not store_result.scalars().first():
            raise HTTPException(status_code=400, detail="Invalid store_id for your company")

    # 1. Create Staff profile ONLY (Personal Details + Password)
    company_id = current_user.company_id
    staff_data = staff_in.model_dump(exclude={"password"})
    
    # --- Uniqueness Checks ---
    # A. Check Email
    if staff_in.email:
        email_check = await db.execute(select(Staff).where((Staff.email == staff_in.email) & (Staff.company_id == company_id)))
        if email_check.scalars().first():
            raise HTTPException(status_code=400, detail=f"Email '{staff_in.email}' is already registered")
            
    # B. Check WhatsApp
    whatsapp_check = await db.execute(select(Staff).where((Staff.whatsapp_number == staff_in.whatsapp_number) & (Staff.company_id == company_id)))
    if whatsapp_check.scalars().first():
        raise HTTPException(status_code=400, detail=f"WhatsApp number '{staff_in.whatsapp_number}' is already registered")

    # Fetch Company for ID generation config
    company = await db.get(Company, current_user.company_id)
    if not company:
        raise HTTPException(status_code=404, detail="Company not found")

    # --- Staff ID Logic ---
    generation_mode = getattr(company, "staff_id_generation_mode", "Automatic")
    
    if generation_mode == "Automatic":
        prefix = company.staff_id_prefix or "STF"
        start_num = company.staff_id_start_number
        padding = company.staff_id_padding or 4
        suffix = company.staff_id_suffix or ""
        generated_code = f"{prefix}{str(start_num).zfill(padding)}{suffix}"
        staff_data["staff_code"] = generated_code
        company.staff_id_start_number += 1
        db.add(company)
    else:
        # Manual Mode
        if not staff_in.staff_code:
            raise HTTPException(status_code=400, detail="Staff ID is required in Manual mode")
        # Check Uniqueness for Manual ID
        code_check = await db.execute(select(Staff).where((Staff.staff_code == staff_in.staff_code) & (Staff.company_id == company_id)))
        if code_check.scalars().first():
            raise HTTPException(status_code=400, detail=f"Staff ID '{staff_in.staff_code}' is already taken")
        staff_data["staff_code"] = staff_in.staff_code.upper()

    # Ensure status is set
    if not staff_data.get("status"):
        staff_data["status"] = "active"

    new_staff = Staff(
        **staff_data,
        company_id=company_id,
        user_id=current_user.user_id,
        password_hash=get_password_hash(staff_in.password)
    )
    db.add(new_staff)
    await db.commit()
    return new_staff


@router.get("/{staff_id}", response_model=StaffOut)
async def get_staff(
    staff_id: uuid.UUID,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_supervisor)
) -> Any:
    result = await db.execute(
        select(Staff).where(
            (Staff.staff_id == staff_id) & (Staff.company_id == current_user.company_id)
        )
    )
    staff = result.scalars().first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    return staff

@router.put("/{staff_id}", response_model=StaffOut)
async def update_staff(
    staff_id: uuid.UUID,
    staff_in: StaffUpdate,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_admin)
) -> Any:
    result = await db.execute(
        select(Staff).where(
            (Staff.staff_id == staff_id) & (Staff.company_id == current_user.company_id)
        )
    )
    staff = result.scalars().first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
    
    update_data = staff_in.model_dump(exclude_unset=True)

    # 1. Uniqueness Checks
    if "email" in update_data and update_data["email"]:
        email_check = await db.execute(select(Staff).where(
            (Staff.email == update_data["email"]) & 
            (Staff.company_id == current_user.company_id) & 
            (Staff.staff_id != staff_id)
        ))
        if email_check.scalars().first():
            raise HTTPException(status_code=400, detail="Email is already used by another employee")

    if "whatsapp_number" in update_data:
        wa_check = await db.execute(select(Staff).where(
            (Staff.whatsapp_number == update_data["whatsapp_number"]) & 
            (Staff.company_id == current_user.company_id) & 
            (Staff.staff_id != staff_id)
        ))
        if wa_check.scalars().first():
            raise HTTPException(status_code=400, detail="WhatsApp number is already used by another employee")

    if "staff_code" in update_data and update_data["staff_code"]:
        code_check = await db.execute(select(Staff).where(
            (Staff.staff_code == update_data["staff_code"].upper()) & 
            (Staff.company_id == current_user.company_id) & 
            (Staff.staff_id != staff_id)
        ))
        if code_check.scalars().first():
            raise HTTPException(status_code=400, detail="Staff ID is already taken")
        update_data["staff_code"] = update_data["staff_code"].upper()

    # 2. Handle store_id verification if provided
    if "store_id" in update_data and update_data["store_id"] is not None:
        store_result = await db.execute(
            select(Store).where(
                (Store.store_id == update_data["store_id"]) & (Store.company_id == current_user.company_id)
            )
        )
        if not store_result.scalars().first():
            raise HTTPException(status_code=400, detail="Invalid store_id for your company")

    # 3. Handle password hashing if provided
    if "password" in update_data:
        from app.core.security import get_password_hash
        staff.password_hash = get_password_hash(update_data["password"])
        del update_data["password"]

    # 4. Update fields
    for field, value in update_data.items():
        if hasattr(staff, field):
            setattr(staff, field, value)
    
    db.add(staff)
    await db.commit()
    
    return staff


@router.get("/import/template")
async def get_import_template(
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_admin)
) -> Any:
    # 1. Fetch available stores for this company
    store_result = await db.execute(select(Store).where(Store.company_id == current_user.company_id))
    store_names = [s.store_name for s in store_result.scalars().all()]
    if not store_names:
        store_names = ["Main Store (No Stores Found)"]
        
    # 2. Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Import Template"
    
    # Define columns
    columns = ["Name", "Email", "WhatsApp Number", "Password", "Role", "Store Name"]
    
    # 0. Fetch company to check generation mode
    company_q = await db.execute(select(Company).where(Company.company_id == current_user.company_id))
    company = company_q.scalars().first()
    generation_mode = getattr(company, "staff_id_generation_mode", "Automatic")
    
    if generation_mode == "Manual":
        columns.insert(0, "Staff ID")
        
    ws.append(columns)
        # Stylize headers and columns
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F764E", end_color="0F764E", fill_type="solid") # Emerald Green
    
    # Set Column Widths
    column_widths = [20, 30, 20, 15, 15, 25]
    if generation_mode == "Manual":
        column_widths.insert(0, 15)
        
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    # Apply header style
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Set WhatsApp Number column (C) as TEXT to prevent scientific notation (1.23E+09)
    # Apply to top 100 rows
    for row in range(2, 102):
        ws.cell(row=row, column=3).number_format = '@' # Text format
        
    # Add a sample row
    sample_row = ["John Doe", "john.doe@example.com", "+1234567890", "Pass@123", "Staff", store_names[0]]
    if generation_mode == "Manual":
        sample_row.insert(0, "STF1001")
    ws.append(sample_row)
    
    # --- Data Validation (Dropdowns) ---
    
    # A. Roles Dropdown
    roles_list = '"Staff,Supervisor"'
    dv_role = DataValidation(type="list", formula1=roles_list, allow_blank=True)
    ws.add_data_validation(dv_role)
    # Apply to Role column (E or F) for next 100 rows
    role_col_idx = 6 if generation_mode == "Manual" else 5
    role_col_letter = chr(64 + role_col_idx)
    dv_role.add(f"{role_col_letter}2:{role_col_letter}102")
    
    # B. Store Names Dropdown
    store_col_idx = 7 if generation_mode == "Manual" else 6
    store_col_letter = chr(64 + store_col_idx)
    
    if len(",".join(store_names)) < 240:
        stores_list = f'"{", ".join(store_names)}"'
        dv_store = DataValidation(type="list", formula1=stores_list, allow_blank=True)
        ws.add_data_validation(dv_store)
        dv_store.add(f"{store_col_letter}2:{store_col_letter}102")
    else:
        # Long store list - use a hidden sheet
        list_ws = wb.create_sheet("HiddenData")
        list_ws.sheet_state = 'hidden'
        for i, name in enumerate(store_names, 1):
            list_ws.cell(row=i, column=1).value = name
        
        dv_store = DataValidation(type="list", formula1=f"HiddenData!$A$1:$A${len(store_names)}", allow_blank=True)
        ws.add_data_validation(dv_store)
        dv_store.add(f"{store_col_letter}2:{store_col_letter}102")

    # 3. Stream back as Excel
    output = io.BytesIO()
    wb.save(output)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=staff_import_template.xlsx"}
    )

@router.post("/import", response_model=dict)
async def import_staff(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_admin)
) -> Any:
    # 0. Validate File Type (for safety)
    allowed_mime_types = [
        "text/csv", 
        "application/vnd.ms-excel", 
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ]
    # Note: Sometimes browser sends generic octet-stream, so we also check extension
    file_ext = Path(file.filename).suffix.lower()
    if file.content_type not in allowed_mime_types and file_ext not in ['.csv', '.xls', '.xlsx']:
        raise HTTPException(status_code=400, detail="Invalid file type. Please upload a CSV or Excel file.")

    # 1. Read file into DataFrame
    contents = await file.read()
    file_ext = Path(file.filename).suffix.lower()
    
    try:
        if file_ext == '.csv':
            df = pd.read_csv(io.BytesIO(contents))
        elif file_ext in ('.xls', '.xlsx'):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use CSV or Excel.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    # 2. Preparation
    from app.core.security import get_password_hash
    
    company_id = current_user.company_id
    
    # Fetch all stores for this company for quick lookup
    store_result = await db.execute(select(Store).where(Store.company_id == company_id))
    stores_map = {s.store_name.lower(): s.store_id for s in store_result.scalars().all()}
    
    # Fetch company for ID generation
    company_q = await db.execute(select(Company).where(Company.company_id == company_id))
    company = company_q.scalars().first()
    
    if not company:
        raise HTTPException(status_code=400, detail="Company not found")

    prefix = company.staff_id_prefix or "STF"
    start_num = company.staff_id_start_number
    padding = company.staff_id_padding or 4
    suffix = company.staff_id_suffix or ""
    
    success_count = 0
    errors = []
    
    # 3. Process Rows
    for index, row in df.iterrows():
        try:
            name = str(row.get("Name", "")).strip()
            email = str(row.get("Email", "")).strip().lower()
            whatsapp = str(row.get("WhatsApp Number", "")).strip()
            password = str(row.get("Password", "Pass@123")).strip()
            role = str(row.get("Role", "Staff")).strip()
            store_name = str(row.get("Store Name", "")).strip()
            
            # Sanitization & Manual Validation for Import
            whatsapp = re.sub(r'[^\d+]', '', whatsapp)
            
            if not name or not email:
                errors.append(f"Row {index + 1}: Name and Email are required")
                continue
            
            if len(name) < 2 or len(name) > 100:
                errors.append(f"Row {index + 1}: Name must be between 2 and 100 characters")
                continue

            if len(email) > 254 or not re.match(r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$", email):
                errors.append(f"Row {index + 1}: Invalid email address")
                continue
            
            if password and (len(password) < 8 or len(password) > 64):
                errors.append(f"Row {index + 1}: Password must be between 8 and 64 characters")
                continue
                
            # Check unique email in company
            email_check = await db.execute(
                select(Staff).where((Staff.email == email) & (Staff.company_id == company_id))
            )
            if email_check.scalars().first():
                errors.append(f"Row {index + 1}: Email '{email}' already registered for this company")
                continue

            # Resolve Store ID
            target_store_id = stores_map.get(store_name.lower())
            
            # Generate or Parse ID
            generation_mode = getattr(company, "staff_id_generation_mode", "Automatic")
            if generation_mode == "Manual":
                staff_code = str(row.get("Staff ID", "")).strip().upper()
                if not staff_code or staff_code == "NAN":
                    errors.append(f"Row {index + 1}: Staff ID is required in Manual mode")
                    continue
                # Validate manual ID format
                if not re.match(r"^[A-Z0-9_-]+$", staff_code):
                    errors.append(f"Row {index + 1}: Staff ID '{staff_code}' contains invalid characters")
                    continue
                # Check uniqueness in current company
                code_check = await db.execute(select(Staff).where(Staff.staff_code == staff_code))
                if code_check.scalars().first():
                    errors.append(f"Row {index + 1}: Staff ID '{staff_code}' already exists")
                    continue
            else:
                staff_code = f"{prefix}{str(start_num).zfill(padding)}{suffix}"
                start_num += 1
            
            new_staff = Staff(
                name=name,
                email=email,
                whatsapp_number=whatsapp,
                role=role,
                store_id=target_store_id,
                staff_code=staff_code,
                status="active",
                company_id=company_id,
                user_id=current_user.user_id,
                password_hash=get_password_hash(password)
            )
            db.add(new_staff)
            success_count += 1
            
        except Exception as e:
            errors.append(f"Row {index + 1}: {str(e)}")

    # 4. Update company start number & Commit
    company.staff_id_start_number = start_num
    db.add(company)
    
    await db.commit()
    
    return {
        "success": success_count,
        "total": len(df),
        "errors": errors
    }

@router.post("/{staff_id}/photo", response_model=StaffOut)
async def upload_staff_photo(
    staff_id: uuid.UUID,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_admin)
) -> Any:
    # 1. Fetch staff
    result = await db.execute(
        select(Staff).where(
            (Staff.staff_id == staff_id) & (Staff.company_id == current_user.company_id)
        )
    )
    staff = result.scalars().first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
        
    # 2. Validate MIME type
    valid_types = ['image/jpeg', 'image/png', 'image/webp']
    if file.content_type not in valid_types:
        raise HTTPException(status_code=400, detail="Invalid file type. Only JPEG, PNG and WebP are allowed.")

    # 3. Save file
    upload_dir = Path("static/uploads")
    upload_dir.mkdir(parents=True, exist_ok=True)
    
    file_extension = Path(file.filename).suffix
    file_name = f"staff_{staff_id}{file_extension}"
    file_path = upload_dir / file_name
    
    with file_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    # 3. Update photo_url
    staff.photo_url = f"/static/uploads/{file_name}"
    
    await db.commit()
    return staff

@router.delete("/{staff_id}")
async def delete_staff(
    staff_id: uuid.UUID,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_admin)
) -> Any:
    # 1. Fetch staff
    result = await db.execute(
        select(Staff).where(
            (Staff.staff_id == staff_id) & (Staff.company_id == current_user.company_id)
        )
    )
    staff = result.scalars().first()
    if not staff:
        raise HTTPException(status_code=404, detail="Staff member not found")
        
    # 2. Delete staff 
    await db.delete(staff)
    await db.commit()
    
    return {"status": "success", "message": "Staff member deleted successfully"}
