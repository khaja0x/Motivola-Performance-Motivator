from typing import Any, List
import uuid
from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.core.database import get_db
from app.api.deps import get_current_user, get_tenant_db
from app.models.models import Store, User, Staff
from app.schemas.store import StoreCreate, StoreUpdate, StoreOut
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

router = APIRouter()

@router.get("/", response_model=List[StoreOut])
async def list_stores(
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    # 1. Fetch all stores for this company
    result = await db.execute(
        select(Store).where(Store.company_id == current_user.company_id)
    )
    stores = result.scalars().all()
    
    # 2. For each store, fetch aggregates (staff count and manager)
    store_outs = []
    for store in stores:
        # Count staff
        staff_res = await db.execute(
            select(Staff).where(Staff.store_id == store.store_id)
        )
        staff_list = staff_res.scalars().all()
        
        # Find manager (Supervisor only)
        manager = next((s.name for s in staff_list if s.role.lower() == 'supervisor'), None)
            
        store_out = StoreOut(
            store_id=store.store_id,
            store_name=store.store_name,
            store_code=store.store_code,
            location=store.location,
            staff_count=len(staff_list),
            manager_name=manager
        )
        store_outs.append(store_out)
        
    return store_outs

@router.post("/", response_model=StoreOut)
async def create_store(
    store_in: StoreCreate,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    # 1. Check if store name already exists for this company
    exist_result = await db.execute(
        select(Store).where(
            (Store.store_name == store_in.store_name) & 
            (Store.company_id == current_user.company_id)
        )
    )
    if exist_result.scalars().first():
        raise HTTPException(status_code=400, detail=f"Store '{store_in.store_name}' already exists")

    # 2. Create store
    new_store = Store(
        **store_in.model_dump(),
        company_id=current_user.company_id
    )
    db.add(new_store)
    await db.commit()
    return new_store

@router.get("/{store_id}", response_model=StoreOut)
async def get_store(
    store_id: uuid.UUID,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(
        select(Store).where(
            (Store.store_id == store_id) & (Store.company_id == current_user.company_id)
        )
    )
    store = result.scalars().first()
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    return store

@router.put("/{store_id}", response_model=StoreOut)
async def update_store(
    store_id: uuid.UUID,
    store_in: StoreUpdate,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(
        select(Store).where(
            (Store.store_id == store_id) & (Store.company_id == current_user.company_id)
        )
    )
    store = result.scalars().first()
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    
    # 2. Check if name is being changed to an existing store name
    update_data = store_in.model_dump(exclude_unset=True)
    if "store_name" in update_data:
        exist_result = await db.execute(
            select(Store).where(
                (Store.store_name == update_data["store_name"]) & 
                (Store.company_id == current_user.company_id) & 
                (Store.store_id != store_id)
            )
        )
        if exist_result.scalars().first():
            raise HTTPException(status_code=400, detail=f"Another store named '{update_data['store_name']}' already exists")

    # 3. Apply updates
    for field, value in update_data.items():
        setattr(store, field, value)
    
    await db.commit()
    return store

@router.delete("/{store_id}")
async def delete_store(
    store_id: uuid.UUID,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(
        select(Store).where(
            (Store.store_id == store_id) & (Store.company_id == current_user.company_id)
        )
    )
    store = result.scalars().first()
    if not store:
        raise HTTPException(status_code=404, detail="Store not found")
    
    await db.delete(store)
    await db.commit()
    return {"status": "success", "message": "Store deleted successfully"}

@router.get("/import/template")
async def get_store_import_template(
    current_user: User = Depends(get_current_user)
) -> Any:
    # 1. Create Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Store Import Template"
    
    # Define columns
    columns = ["Store Name", "Store Code", "Location/City"]
    ws.append(columns)
    
    # Stylize headers and columns
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Slate 800
    
    # Set Column Widths
    column_widths = [30, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64+i)].width = width

    # Apply header style
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    # Add a sample row
    ws.append(["Main Branch", "MB001", "New York, NY"])
    
    # 3. Return as Excel
    output = io.BytesIO()
    wb.save(output)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=store_import_template.xlsx"}
    )

@router.post("/import", response_model=dict)
async def import_stores(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    # 1. Read file
    contents = await file.read()
    import os
    file_ext = os.path.splitext(file.filename)[1].lower()
    
    try:
        if file_ext == '.csv':
            df = pd.read_csv(io.BytesIO(contents))
        elif file_ext in ('.xls', '.xlsx'):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format. Please use CSV or Excel.")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    success_count = 0
    errors = []
    
    # 2. Process Rows
    for index, row in df.iterrows():
        try:
            store_name = str(row.get("Store Name", "")).strip()
            store_code = str(row.get("Store Code", "")).strip()
            location = str(row.get("Location/City", "")).strip()
            
            if not store_name:
                errors.append(f"Row {index + 2}: Store Name is required")
                continue

            # Check for duplicates in DB for this company
            exist_result = await db.execute(
                select(Store).where(
                    (Store.store_name == store_name) & 
                    (Store.company_id == current_user.company_id)
                )
            )
            if exist_result.scalars().first():
                errors.append(f"Row {index + 2}: Store '{store_name}' already exists")
                continue
                
            new_store = Store(
                store_name=store_name,
                store_code=store_code if store_code else None,
                location=location if location else None,
                company_id=current_user.company_id
            )
            db.add(new_store)
            success_count += 1
            
        except Exception as e:
            errors.append(f"Row {index + 2}: {str(e)}")

    await db.commit()
    
    return {
        "success": success_count,
        "total": len(df),
        "errors": errors
    }
