from typing import Any, List, Optional
import uuid
from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from app.api.deps import get_tenant_db, get_current_user
from app.models.models import Target, User
from app.schemas.target import TargetCreate, TargetUpdate, TargetOut

router = APIRouter()

@router.get("/", response_model=List[TargetOut])
async def list_targets(
    period_type: str = Query("Monthly"),
    month: Optional[int] = Query(None),
    week: Optional[int] = Query(None),
    day: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    query = select(Target).where(Target.period_type == period_type)
    
    if year:
        query = query.where(Target.year == year)
    if month:
        query = query.where(Target.month == month)
    if week:
        query = query.where(Target.week == week)
    if day:
        query = query.where(Target.day == day)
        
    query = query.where(Target.company_id == current_user.company_id)
    
    result = await db.execute(query)
    return result.scalars().all()

@router.post("/", response_model=TargetOut)
async def create_target(
    target_in: TargetCreate,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    # Check if target already exists for exactly this period
    check_query = select(Target).where(
        (Target.entity_type == target_in.entity_type) &
        (Target.entity_id == target_in.entity_id) &
        (Target.period_type == target_in.period_type) &
        (Target.year == target_in.year)
    )
    
    if target_in.month:
        check_query = check_query.where(Target.month == target_in.month)
    if target_in.week:
        check_query = check_query.where(Target.week == target_in.week)
    if target_in.day:
        check_query = check_query.where(Target.day == target_in.day)

    result = await db.execute(check_query)
    if result.scalars().first():
        raise HTTPException(status_code=400, detail="Target already exists for this period")

    new_target = Target(
        **target_in.model_dump(),
        company_id=current_user.company_id
    )
    db.add(new_target)
    await db.commit()
    return new_target


@router.put("/{target_id}", response_model=TargetOut)
async def update_target(
    target_id: uuid.UUID,
    target_in: TargetUpdate,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(select(Target).where(Target.target_id == target_id))
    target = result.scalars().first()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    
    update_data = target_in.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(target, field, value)
    
    await db.commit()
    return target

@router.delete("/{target_id}")
async def delete_target(
    target_id: uuid.UUID,
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    result = await db.execute(select(Target).where(Target.target_id == target_id))
    target = result.scalars().first()
    if not target:
        raise HTTPException(status_code=404, detail="Target not found")
    
    await db.delete(target)
    await db.commit()
    return {"status": "success"}

# --- Bulk Import Support ---

from fastapi import Response, UploadFile, File
import io
import pandas as pd
from openpyxl import Workbook
from app.models.models import Staff
import re

@router.get("/import/template")
async def get_target_import_template(
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    # Fetch staff for the company to include in template
    staff_result = await db.execute(select(Staff).where(Staff.company_id == current_user.company_id))
    staff_list = staff_result.scalars().all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Target Import Template"
    
    headers = ["Staff Code", "Staff Name", "Period Type", "Year", "Month", "Week", "Day", "Target Amount", "Target Quantity"]
    ws.append(headers)
    
    # Styling
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Add staff as examples
    for s in staff_list:
        ws.append([s.staff_code, s.name, "Monthly", 2026, 4, None, None, 50000, 100])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=targets_import_template.xlsx"}
    )

@router.post("/import")
async def import_targets(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    contents = await file.read()
    file_ext = file.filename.split('.')[-1].lower()
    
    try:
        if file_ext == 'csv':
            df = pd.read_csv(io.BytesIO(contents))
        elif file_ext in ('xls', 'xlsx'):
            df = pd.read_excel(io.BytesIO(contents))
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    success_count = 0
    errors = []
    
    # Map staff_code to staff_id
    staff_res = await db.execute(select(Staff).where(Staff.company_id == current_user.company_id))
    staff_map = {s.staff_code: s.staff_id for s in staff_res.scalars().all()}

    for index, row in df.iterrows():
        try:
            staff_code = str(row.get("Staff Code", "")).strip()
            period_type = str(row.get("Period Type", "Monthly")).strip()
            year = int(row.get("Year", 2024))
            month = row.get("Month")
            week = row.get("Week")
            day = row.get("Day")
            amount = float(row.get("Target Amount", 0))
            quantity = int(row.get("Target Quantity", 0))

            staff_id = staff_map.get(staff_code)
            if not staff_id:
                errors.append(f"Row {index+2}: Staff code {staff_code} not found")
                continue

            # Check if exists
            q = select(Target).where(
                (Target.entity_id == staff_id) &
                (Target.period_type == period_type) &
                (Target.year == year)
            )
            if month and not pd.isna(month): q = q.where(Target.month == int(month))
            if week and not pd.isna(week): q = q.where(Target.week == int(week))
            if day and not pd.isna(day): q = q.where(Target.day == int(day))
            
            res = await db.execute(q)
            target = res.scalars().first()

            if target:
                target.target_amount = amount
                target.target_quantity = quantity
            else:
                new_target = Target(
                    company_id=current_user.company_id,
                    entity_type="Staff",
                    entity_id=staff_id,
                    period_type=period_type,
                    year=year,
                    month=int(month) if month and not pd.isna(month) else None,
                    week=int(week) if week and not pd.isna(week) else None,
                    day=int(day) if day and not pd.isna(day) else None,
                    target_amount=amount,
                    target_quantity=quantity
                )
                db.add(new_target)
            
            success_count += 1
        except Exception as e:
            errors.append(f"Row {index+2}: {str(e)}")

    await db.commit()
    return {"success": success_count, "errors": errors}

@router.post("/bulk", response_model=List[TargetOut])
async def bulk_upsert_targets(
    targets_in: List[TargetCreate],
    db: AsyncSession = Depends(get_tenant_db),
    current_user: User = Depends(get_current_user)
) -> Any:
    results = []
    for t_in in targets_in:
        # Check if target already exists
        check_query = select(Target).where(
            (Target.entity_type == t_in.entity_type) &
            (Target.entity_id == t_in.entity_id) &
            (Target.period_type == t_in.period_type) &
            (Target.year == t_in.year) &
            (Target.company_id == current_user.company_id)
        )
        
        if t_in.month:
            check_query = check_query.where(Target.month == t_in.month)
        if t_in.week:
            check_query = check_query.where(Target.week == t_in.week)
        if t_in.day:
            check_query = check_query.where(Target.day == t_in.day)

        res = await db.execute(check_query)
        existing = res.scalars().first()

        if existing:
            # Update
            update_data = t_in.model_dump(exclude_unset=True)
            for field, value in update_data.items():
                setattr(existing, field, value)
            results.append(existing)
        else:
            # Create
            new_target = Target(
                **t_in.model_dump(),
                company_id=current_user.company_id
            )
            db.add(new_target)
            results.append(new_target)

    await db.commit()
    return results

